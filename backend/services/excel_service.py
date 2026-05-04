"""
Excel export service using openpyxl.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _style_header(ws, headers, row=1):
    """Apply styling to header row."""
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)


def export_transactions(transactions: list) -> io.BytesIO:
    """Generate Excel file from transactions data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"

    headers = ["No", "Tanggal", "Produk", "Tipe", "Jumlah", "Total Harga", "Catatan"]
    _style_header(ws, headers)

    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    total_penjualan = 0
    total_pembelian = 0

    for i, tx in enumerate(transactions, 1):
        row = i + 1

        tx_type = tx.get("type")
        total_price = tx.get("total_price", 0)

        if tx_type == "sale":
            tipe_label = "Penjualan"
            signed_total_price = total_price
            total_penjualan += total_price
        else:
            tipe_label = "Pembelian"
            signed_total_price = -total_price
            total_pembelian += total_price

        values = [
            i,
            tx.get("created_at", "")[:16].replace("T", " "),
            tx.get("product_name", ""),
            tipe_label,
            tx.get("quantity", 0),
            signed_total_price,
            tx.get("note", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")

            if col == 6:
                cell.number_format = '+#,##0;-#,##0;0'

    # Keuntungan bersih di bawah tabel
    summary_start_row = len(transactions) + 2
    keuntungan_bersih = total_penjualan - total_pembelian

    summary_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    summary_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    label_cell = ws.cell(row=summary_start_row, column=5, value="Keuntungan Bersih")
    value_cell = ws.cell(row=summary_start_row, column=6, value=keuntungan_bersih)

    label_cell.font = summary_font
    value_cell.font = summary_font

    label_cell.fill = summary_fill
    value_cell.fill = summary_fill

    label_cell.border = thin_border
    value_cell.border = thin_border

    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    value_cell.alignment = Alignment(horizontal="right", vertical="center")

    value_cell.number_format = '#,##0'

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_products(products: list) -> io.BytesIO:
    """Generate Excel file from products data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Produk"

    headers = ["No", "Nama Produk", "SKU", "Kategori", "Stok", "Harga"]
    _style_header(ws, headers)

    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for i, p in enumerate(products, 1):
        row = i + 1

        values = [
            i,
            p.get("name", ""),
            p.get("sku", ""),
            p.get("category", ""),
            p.get("stock", 0),
            p.get("price", 0),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")

            if col == 6:
                cell.number_format = '#,##0'

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output